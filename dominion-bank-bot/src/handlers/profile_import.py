"""
The Phantom Bot - Profile Import/Export Handler
/plantilla_perfiles - Generate blank Excel template for Google Sheets
/exportar_perfiles - Export all profiles to Excel
/importar_perfiles - Import profiles from Excel
"""
import logging
import tempfile
from pathlib import Path
from telegram import Update
from telegram.ext import ContextTypes
from sqlalchemy import select

from src.database.connection import get_session
from src.database.repositories import UserRepository
from src.database.repositories.profile import ProfileRepository, UserSettingsRepository
from src.database.models import (
    User, Profile, UserSettings, Kink, UserKink, UserLimit,
    MainRole, ExperienceLevel, PrivacyLevel, LimitType
)
from src.utils.helpers import is_admin

logger = logging.getLogger(__name__)


# =============================================================================
# EXCEL STYLING HELPERS
# =============================================================================

def _style_header_row(sheet, headers, header_fill, header_font):
    """Apply styling to header row."""
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font


def _add_data_validation(sheet, col_letter, options, max_row=1000):
    """Add dropdown validation to a column."""
    from openpyxl.worksheet.datavalidation import DataValidation
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(options)}"',
        allow_blank=True
    )
    dv.error = "Valor no válido"
    dv.errorTitle = "Error"
    dv.prompt = "Selecciona una opción"
    dv.promptTitle = "Opciones"
    sheet.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{max_row}")


def _create_instructions_sheet(workbook):
    """Create an instructions sheet for the template."""
    sheet = workbook.create_sheet("Instrucciones", 0)

    instructions = [
        ("INSTRUCCIONES PARA RELLENAR EL ARCHIVO", ""),
        ("", ""),
        ("1. HOJA 'Perfiles'", ""),
        ("   - username: @usuario de Telegram (obligatorio)", ""),
        ("   - display_name: Nombre que se muestra en el perfil", ""),
        ("   - pronouns: él/ella/elle u otro", ""),
        ("   - age: Edad (número)", ""),
        ("   - show_age: exact (mostrar edad) / range (rango) / hidden (oculto)", ""),
        ("   - location: Ubicación/Ciudad", ""),
        ("   - main_role: dom / sub / switch", ""),
        ("   - sub_roles: Roles secundarios separados por coma", ""),
        ("   - experience_level: principiante / intermedio / avanzado / experto", ""),
        ("   - bio: Biografía/descripción personal", ""),
        ("   - looking_for: Qué busca", ""),
        ("   - availability: Disponibilidad horaria", ""),
        ("", ""),
        ("2. HOJA 'Configuracion'", ""),
        ("   - username: @usuario de Telegram (obligatorio)", ""),
        ("   - privacy_level: public / members / verified / private", ""),
        ("   - notify_transfers: TRUE / FALSE", ""),
        ("   - notify_mentions: TRUE / FALSE", ""),
        ("   - notify_bdsm: TRUE / FALSE", ""),
        ("", ""),
        ("3. HOJA 'Limites'", ""),
        ("   - username: @usuario de Telegram (obligatorio)", ""),
        ("   - limit_type: hard (límite duro) / soft (límite blando)", ""),
        ("   - description: Descripción del límite", ""),
        ("", ""),
        ("4. HOJA 'Kinks' (Referencia)", ""),
        ("   - Lista de kinks disponibles para referencia", ""),
        ("", ""),
        ("5. HOJA 'UserKinks'", ""),
        ("   - username: @usuario de Telegram (obligatorio)", ""),
        ("   - kink_name: Nombre del kink (de la lista de referencia)", ""),
        ("   - level: 1-3 (nivel de interés)", ""),
        ("   - direction: give / receive / both", ""),
        ("   - curious: TRUE / FALSE (si es curioso por explorar)", ""),
        ("", ""),
        ("NOTAS IMPORTANTES:", ""),
        ("- Descarga este archivo y súbelo a Google Sheets", ""),
        ("- Los usuarios pueden rellenar sus datos online", ""),
        ("- Descarga como .xlsx y usa /importar_perfiles", ""),
        ("- Los usuarios existentes serán actualizados", ""),
        ("- Los usuarios nuevos serán creados sin Telegram ID", ""),
    ]

    for row_num, (text, _) in enumerate(instructions, 1):
        sheet.cell(row=row_num, column=1, value=text)

    sheet.column_dimensions["A"].width = 80


# =============================================================================
# TEMPLATE COMMAND
# =============================================================================

async def plantilla_perfiles_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /plantilla_perfiles - Generate blank Excel template."""
    if not update.effective_user or not update.message:
        return

    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Solo los administradores pueden generar plantillas.")
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        await update.message.reply_text(
            "❌ Error: La librería openpyxl no está instalada.\n"
            "Ejecuta: pip install openpyxl"
        )
        return

    await update.message.reply_text("📝 Generando plantilla de perfiles...")

    try:
        workbook = openpyxl.Workbook()

        # Remove default sheet
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        # Styling
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        # 1. Create Instructions sheet
        _create_instructions_sheet(workbook)

        # 2. Profiles sheet
        profiles_sheet = workbook.create_sheet("Perfiles")
        profile_headers = [
            "username", "display_name", "pronouns", "age", "show_age",
            "location", "main_role", "sub_roles", "experience_level",
            "bio", "looking_for", "availability"
        ]
        _style_header_row(profiles_sheet, profile_headers, header_fill, header_font)

        # Add dropdowns for enum fields
        _add_data_validation(profiles_sheet, "E", ["exact", "range", "hidden"])  # show_age
        _add_data_validation(profiles_sheet, "G", ["dom", "sub", "switch"])  # main_role
        _add_data_validation(profiles_sheet, "I", ["principiante", "intermedio", "avanzado", "experto"])  # experience

        for col in range(1, len(profile_headers) + 1):
            profiles_sheet.column_dimensions[get_column_letter(col)].width = 15
        profiles_sheet.column_dimensions["J"].width = 40  # bio
        profiles_sheet.column_dimensions["K"].width = 30  # looking_for

        # 3. Settings sheet
        settings_sheet = workbook.create_sheet("Configuracion")
        settings_headers = [
            "username", "privacy_level", "notify_transfers",
            "notify_mentions", "notify_bdsm"
        ]
        _style_header_row(settings_sheet, settings_headers, header_fill, header_font)
        _add_data_validation(settings_sheet, "B", ["public", "members", "verified", "private"])
        _add_data_validation(settings_sheet, "C", ["TRUE", "FALSE"])
        _add_data_validation(settings_sheet, "D", ["TRUE", "FALSE"])
        _add_data_validation(settings_sheet, "E", ["TRUE", "FALSE"])

        for col in range(1, len(settings_headers) + 1):
            settings_sheet.column_dimensions[get_column_letter(col)].width = 18

        # 4. Limits sheet
        limits_sheet = workbook.create_sheet("Limites")
        limits_headers = ["username", "limit_type", "description"]
        _style_header_row(limits_sheet, limits_headers, header_fill, header_font)
        _add_data_validation(limits_sheet, "B", ["hard", "soft"])

        limits_sheet.column_dimensions["A"].width = 15
        limits_sheet.column_dimensions["B"].width = 12
        limits_sheet.column_dimensions["C"].width = 50

        # 5. Kinks reference sheet
        kinks_sheet = workbook.create_sheet("Kinks")
        kinks_headers = ["id", "name", "category", "description"]
        _style_header_row(kinks_sheet, kinks_headers, header_fill, header_font)

        # Add existing kinks from database
        async with get_session() as session:
            result = await session.execute(select(Kink).order_by(Kink.category, Kink.name))
            kinks = result.scalars().all()

            for row_num, kink in enumerate(kinks, start=2):
                kinks_sheet.cell(row=row_num, column=1, value=kink.id)
                kinks_sheet.cell(row=row_num, column=2, value=kink.name)
                kinks_sheet.cell(row=row_num, column=3, value=kink.category)
                kinks_sheet.cell(row=row_num, column=4, value=kink.description or "")

        for col in range(1, 5):
            kinks_sheet.column_dimensions[get_column_letter(col)].width = 20
        kinks_sheet.column_dimensions["D"].width = 40

        # 6. UserKinks sheet
        user_kinks_sheet = workbook.create_sheet("UserKinks")
        user_kinks_headers = ["username", "kink_name", "level", "direction", "curious"]
        _style_header_row(user_kinks_sheet, user_kinks_headers, header_fill, header_font)
        _add_data_validation(user_kinks_sheet, "C", ["1", "2", "3"])
        _add_data_validation(user_kinks_sheet, "D", ["give", "receive", "both"])
        _add_data_validation(user_kinks_sheet, "E", ["TRUE", "FALSE"])

        for col in range(1, len(user_kinks_headers) + 1):
            user_kinks_sheet.column_dimensions[get_column_letter(col)].width = 15

        # Save and send
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            workbook.save(tmp.name)
            tmp_path = Path(tmp.name)

        with open(tmp_path, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename="phantom_plantilla_perfiles.xlsx",
                caption=(
                    "📝 **Plantilla de Perfiles**\n\n"
                    "1. Sube este archivo a Google Sheets\n"
                    "2. Comparte el enlace con los usuarios\n"
                    "3. Cada usuario rellena su fila\n"
                    "4. Descarga como .xlsx\n"
                    "5. Usa /importar_perfiles para cargar los datos"
                )
            )

        tmp_path.unlink(missing_ok=True)
        logger.info(f"Profile template generated by {update.effective_user.id}")

    except Exception as e:
        logger.error(f"Template generation error: {e}")
        await update.message.reply_text(f"❌ Error al generar plantilla: {str(e)}")


# =============================================================================
# EXPORT COMMAND
# =============================================================================

async def exportar_perfiles_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /exportar_perfiles - Export all profiles to Excel."""
    if not update.effective_user or not update.message:
        return

    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Solo los administradores pueden exportar perfiles.")
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        await update.message.reply_text(
            "❌ Error: La librería openpyxl no está instalada."
        )
        return

    await update.message.reply_text("📊 Exportando perfiles...")

    try:
        workbook = openpyxl.Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        async with get_session() as session:
            # 1. Export Users + Profiles
            profiles_sheet = workbook.create_sheet("Perfiles")
            profile_headers = [
                "username", "telegram_id", "balance", "display_name", "pronouns",
                "age", "show_age", "location", "main_role", "sub_roles",
                "experience_level", "bio", "looking_for", "availability", "verified"
            ]
            _style_header_row(profiles_sheet, profile_headers, header_fill, header_font)

            result = await session.execute(
                select(User).outerjoin(Profile).order_by(User.username)
            )
            users = result.scalars().unique().all()

            for row_num, user in enumerate(users, start=2):
                profile = user.profile
                profiles_sheet.cell(row=row_num, column=1, value=user.username or "")
                profiles_sheet.cell(row=row_num, column=2, value=user.telegram_id)
                profiles_sheet.cell(row=row_num, column=3, value=user.balance)

                if profile:
                    profiles_sheet.cell(row=row_num, column=4, value=profile.display_name or "")
                    profiles_sheet.cell(row=row_num, column=5, value=profile.pronouns or "")
                    profiles_sheet.cell(row=row_num, column=6, value=profile.age)
                    profiles_sheet.cell(row=row_num, column=7, value=profile.show_age)
                    profiles_sheet.cell(row=row_num, column=8, value=profile.location or "")
                    profiles_sheet.cell(row=row_num, column=9, value=profile.main_role.value if profile.main_role else "")
                    profiles_sheet.cell(row=row_num, column=10, value=profile.sub_roles or "")
                    profiles_sheet.cell(row=row_num, column=11, value=profile.experience_level.value if profile.experience_level else "")
                    profiles_sheet.cell(row=row_num, column=12, value=profile.bio or "")
                    profiles_sheet.cell(row=row_num, column=13, value=profile.looking_for or "")
                    profiles_sheet.cell(row=row_num, column=14, value=profile.availability or "")
                    profiles_sheet.cell(row=row_num, column=15, value=profile.verified)

            for col in range(1, len(profile_headers) + 1):
                profiles_sheet.column_dimensions[get_column_letter(col)].width = 15
            profiles_sheet.column_dimensions["L"].width = 40  # bio

            # 2. Export Settings
            settings_sheet = workbook.create_sheet("Configuracion")
            settings_headers = [
                "username", "privacy_level", "notify_transfers",
                "notify_mentions", "notify_bdsm"
            ]
            _style_header_row(settings_sheet, settings_headers, header_fill, header_font)

            result = await session.execute(
                select(User, UserSettings).outerjoin(UserSettings).order_by(User.username)
            )
            rows = result.all()

            for row_num, (user, user_settings) in enumerate(rows, start=2):
                settings_sheet.cell(row=row_num, column=1, value=user.username or "")
                if user_settings:
                    settings_sheet.cell(row=row_num, column=2, value=user_settings.privacy_level.value)
                    settings_sheet.cell(row=row_num, column=3, value=user_settings.notify_transfers)
                    settings_sheet.cell(row=row_num, column=4, value=user_settings.notify_mentions)
                    settings_sheet.cell(row=row_num, column=5, value=user_settings.notify_bdsm)

            for col in range(1, len(settings_headers) + 1):
                settings_sheet.column_dimensions[get_column_letter(col)].width = 18

            # 3. Export Limits
            limits_sheet = workbook.create_sheet("Limites")
            limits_headers = ["username", "limit_type", "description"]
            _style_header_row(limits_sheet, limits_headers, header_fill, header_font)

            result = await session.execute(
                select(UserLimit, User)
                .join(Profile, UserLimit.user_id == Profile.user_id)
                .join(User, Profile.user_id == User.id)
                .order_by(User.username)
            )
            limit_rows = result.all()

            for row_num, (limit, user) in enumerate(limit_rows, start=2):
                limits_sheet.cell(row=row_num, column=1, value=user.username or "")
                limits_sheet.cell(row=row_num, column=2, value=limit.limit_type.value)
                limits_sheet.cell(row=row_num, column=3, value=limit.description)

            limits_sheet.column_dimensions["A"].width = 15
            limits_sheet.column_dimensions["B"].width = 12
            limits_sheet.column_dimensions["C"].width = 50

            # 4. Export Kinks reference
            kinks_sheet = workbook.create_sheet("Kinks")
            kinks_headers = ["id", "name", "category", "description"]
            _style_header_row(kinks_sheet, kinks_headers, header_fill, header_font)

            result = await session.execute(select(Kink).order_by(Kink.category, Kink.name))
            kinks = result.scalars().all()

            for row_num, kink in enumerate(kinks, start=2):
                kinks_sheet.cell(row=row_num, column=1, value=kink.id)
                kinks_sheet.cell(row=row_num, column=2, value=kink.name)
                kinks_sheet.cell(row=row_num, column=3, value=kink.category)
                kinks_sheet.cell(row=row_num, column=4, value=kink.description or "")

            for col in range(1, 5):
                kinks_sheet.column_dimensions[get_column_letter(col)].width = 20

            # 5. Export UserKinks
            user_kinks_sheet = workbook.create_sheet("UserKinks")
            user_kinks_headers = ["username", "kink_name", "level", "direction", "curious"]
            _style_header_row(user_kinks_sheet, user_kinks_headers, header_fill, header_font)

            result = await session.execute(
                select(UserKink, User, Kink)
                .join(Profile, UserKink.user_id == Profile.user_id)
                .join(User, Profile.user_id == User.id)
                .join(Kink, UserKink.kink_id == Kink.id)
                .order_by(User.username, Kink.name)
            )
            user_kink_rows = result.all()

            for row_num, (user_kink, user, kink) in enumerate(user_kink_rows, start=2):
                user_kinks_sheet.cell(row=row_num, column=1, value=user.username or "")
                user_kinks_sheet.cell(row=row_num, column=2, value=kink.name)
                user_kinks_sheet.cell(row=row_num, column=3, value=user_kink.level)
                user_kinks_sheet.cell(row=row_num, column=4, value=user_kink.direction)
                user_kinks_sheet.cell(row=row_num, column=5, value=user_kink.curious)

            for col in range(1, len(user_kinks_headers) + 1):
                user_kinks_sheet.column_dimensions[get_column_letter(col)].width = 15

        # Save and send
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            workbook.save(tmp.name)
            tmp_path = Path(tmp.name)

        with open(tmp_path, 'rb') as f:
            await update.message.reply_document(
                document=f,
                filename=f"phantom_perfiles_{len(users)}_usuarios.xlsx",
                caption=f"📊 **Exportación de Perfiles**\n\n✅ {len(users)} usuarios exportados"
            )

        tmp_path.unlink(missing_ok=True)
        logger.info(f"Profiles exported by {update.effective_user.id}: {len(users)} users")

    except Exception as e:
        logger.error(f"Profile export error: {e}")
        await update.message.reply_text(f"❌ Error al exportar perfiles: {str(e)}")


# =============================================================================
# IMPORT COMMAND
# =============================================================================

async def importar_perfiles_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /importar_perfiles - Show import instructions."""
    if not update.effective_user or not update.message:
        return

    if not await is_admin(update.effective_user.id):
        await update.message.reply_text("❌ Solo los administradores pueden importar perfiles.")
        return

    await update.message.reply_text(
        """📥 **Importar Perfiles desde Excel**

Para importar perfiles:

1. Usa /plantilla_perfiles para obtener la plantilla
2. Rellena los datos en Google Sheets
3. Descarga como archivo Excel (.xlsx)
4. Envía el archivo como respuesta a este mensaje

**Hojas procesadas:**
- **Perfiles**: Datos básicos del perfil
- **Configuracion**: Privacidad y notificaciones
- **Limites**: Límites duros y blandos
- **UserKinks**: Kinks de cada usuario

⚠️ Los usuarios existentes serán actualizados.
⚠️ Los usuarios nuevos se crean sin Telegram ID."""
    )


async def handle_profile_excel_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle Excel document uploads for profile import."""
    if not update.effective_user or not update.message or not update.message.document:
        return

    if not await is_admin(update.effective_user.id):
        return

    document = update.message.document

    # Check file extension
    if not document.file_name or not document.file_name.endswith(('.xlsx', '.xls')):
        return  # Let other handlers deal with non-Excel files

    # Check if this is a profile import (check for specific sheets)
    try:
        import openpyxl
    except ImportError:
        await update.message.reply_text("❌ Error: openpyxl no está instalada.")
        return

    try:
        # Download file
        file = await context.bot.get_file(document.file_id)

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            await file.download_to_drive(tmp.name)
            tmp_path = Path(tmp.name)

        workbook = openpyxl.load_workbook(tmp_path, read_only=True)
        sheet_names = workbook.sheetnames

        # Check if this is a profile file (has Perfiles sheet)
        if "Perfiles" not in sheet_names:
            tmp_path.unlink(missing_ok=True)
            return  # Not a profile import file

        workbook.close()

        # Reopen for writing
        workbook = openpyxl.load_workbook(tmp_path)

        await update.message.reply_text("📥 Procesando archivo de perfiles...")

        stats = {
            "profiles_created": 0,
            "profiles_updated": 0,
            "settings_updated": 0,
            "limits_created": 0,
            "user_kinks_created": 0,
            "errors": []
        }

        async with get_session() as session:
            user_repo = UserRepository(session)
            profile_repo = ProfileRepository(session)
            settings_repo = UserSettingsRepository(session)

            # Cache for username -> user_id mapping
            user_cache = {}

            async def get_user_by_username(username: str):
                """Get or create user by username."""
                username = username.strip().lstrip("@")
                if not username:
                    return None

                if username in user_cache:
                    return user_cache[username]

                user = await user_repo.get_by_username(username)
                if not user:
                    # Create placeholder user
                    user = await user_repo.create_placeholder(
                        username=username,
                        first_name=username,
                        balance=0
                    )
                    stats["profiles_created"] += 1

                user_cache[username] = user
                return user

            # 1. Process Profiles sheet
            if "Perfiles" in sheet_names:
                sheet = workbook["Perfiles"]
                headers = [cell.value.lower().strip() if cell.value else "" for cell in sheet[1]]

                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        row_dict = dict(zip(headers, row))
                        username = row_dict.get("username", "")
                        if not username:
                            continue

                        user = await get_user_by_username(username)
                        if not user:
                            continue

                        # Get or create profile
                        profile = await profile_repo.get_or_create(user.id)

                        # Update profile fields
                        update_data = {}
                        if row_dict.get("display_name"):
                            update_data["display_name"] = str(row_dict["display_name"])
                        if row_dict.get("pronouns"):
                            update_data["pronouns"] = str(row_dict["pronouns"])
                        if row_dict.get("age"):
                            try:
                                update_data["age"] = int(row_dict["age"])
                            except (ValueError, TypeError):
                                pass
                        if row_dict.get("show_age"):
                            update_data["show_age"] = str(row_dict["show_age"])
                        if row_dict.get("location"):
                            update_data["location"] = str(row_dict["location"])
                        if row_dict.get("main_role"):
                            role_val = str(row_dict["main_role"]).lower()
                            if role_val in ["dom", "sub", "switch"]:
                                update_data["main_role"] = MainRole(role_val)
                        if row_dict.get("sub_roles"):
                            update_data["sub_roles"] = str(row_dict["sub_roles"])
                        if row_dict.get("experience_level"):
                            exp_val = str(row_dict["experience_level"]).lower()
                            if exp_val in ["principiante", "intermedio", "avanzado", "experto"]:
                                update_data["experience_level"] = ExperienceLevel(exp_val)
                        if row_dict.get("bio"):
                            update_data["bio"] = str(row_dict["bio"])
                        if row_dict.get("looking_for"):
                            update_data["looking_for"] = str(row_dict["looking_for"])
                        if row_dict.get("availability"):
                            update_data["availability"] = str(row_dict["availability"])

                        if update_data:
                            await profile_repo.update(user.id, **update_data)
                            stats["profiles_updated"] += 1

                    except Exception as e:
                        stats["errors"].append(f"Perfiles fila {row_num}: {str(e)}")

            # 2. Process Settings sheet
            if "Configuracion" in sheet_names:
                sheet = workbook["Configuracion"]
                headers = [cell.value.lower().strip() if cell.value else "" for cell in sheet[1]]

                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        row_dict = dict(zip(headers, row))
                        username = row_dict.get("username", "")
                        if not username:
                            continue

                        user = await get_user_by_username(username)
                        if not user:
                            continue

                        settings = await settings_repo.get_or_create(user.id)

                        update_data = {}
                        if row_dict.get("privacy_level"):
                            priv_val = str(row_dict["privacy_level"]).lower()
                            if priv_val in ["public", "members", "verified", "private"]:
                                update_data["privacy_level"] = PrivacyLevel(priv_val)

                        for field in ["notify_transfers", "notify_mentions", "notify_bdsm"]:
                            if row_dict.get(field) is not None:
                                val = row_dict[field]
                                if isinstance(val, bool):
                                    update_data[field] = val
                                elif str(val).upper() in ["TRUE", "1", "YES", "SI", "SÍ"]:
                                    update_data[field] = True
                                elif str(val).upper() in ["FALSE", "0", "NO"]:
                                    update_data[field] = False

                        if update_data:
                            await settings_repo.update(user.id, **update_data)
                            stats["settings_updated"] += 1

                    except Exception as e:
                        stats["errors"].append(f"Configuracion fila {row_num}: {str(e)}")

            # 3. Process Limits sheet
            if "Limites" in sheet_names:
                sheet = workbook["Limites"]
                headers = [cell.value.lower().strip() if cell.value else "" for cell in sheet[1]]

                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        row_dict = dict(zip(headers, row))
                        username = row_dict.get("username", "")
                        if not username or not row_dict.get("description"):
                            continue

                        user = await get_user_by_username(username)
                        if not user:
                            continue

                        # Ensure profile exists
                        await profile_repo.get_or_create(user.id)

                        limit_type_val = str(row_dict.get("limit_type", "soft")).lower()
                        limit_type = LimitType.HARD if limit_type_val == "hard" else LimitType.SOFT

                        limit = UserLimit(
                            user_id=user.id,
                            limit_type=limit_type,
                            description=str(row_dict["description"])
                        )
                        session.add(limit)
                        stats["limits_created"] += 1

                    except Exception as e:
                        stats["errors"].append(f"Limites fila {row_num}: {str(e)}")

            # 4. Process UserKinks sheet
            if "UserKinks" in sheet_names:
                sheet = workbook["UserKinks"]
                headers = [cell.value.lower().strip() if cell.value else "" for cell in sheet[1]]

                # Build kink name -> id cache
                kink_cache = {}
                result = await session.execute(select(Kink))
                for kink in result.scalars():
                    kink_cache[kink.name.lower()] = kink.id

                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        row_dict = dict(zip(headers, row))
                        username = row_dict.get("username", "")
                        kink_name = row_dict.get("kink_name", "")
                        if not username or not kink_name:
                            continue

                        user = await get_user_by_username(username)
                        if not user:
                            continue

                        kink_id = kink_cache.get(kink_name.lower())
                        if not kink_id:
                            stats["errors"].append(f"UserKinks fila {row_num}: Kink '{kink_name}' no encontrado")
                            continue

                        # Ensure profile exists
                        await profile_repo.get_or_create(user.id)

                        level = int(row_dict.get("level", 1))
                        if level < 1:
                            level = 1
                        if level > 3:
                            level = 3

                        direction = str(row_dict.get("direction", "both")).lower()
                        if direction not in ["give", "receive", "both"]:
                            direction = "both"

                        curious_val = row_dict.get("curious", False)
                        if isinstance(curious_val, bool):
                            curious = curious_val
                        else:
                            curious = str(curious_val).upper() in ["TRUE", "1", "YES", "SI", "SÍ"]

                        user_kink = UserKink(
                            user_id=user.id,
                            kink_id=kink_id,
                            level=level,
                            direction=direction,
                            curious=curious
                        )
                        session.add(user_kink)
                        stats["user_kinks_created"] += 1

                    except Exception as e:
                        stats["errors"].append(f"UserKinks fila {row_num}: {str(e)}")

            await session.commit()

        # Clean up
        tmp_path.unlink(missing_ok=True)

        # Build response
        response = f"""📥 **Importación de Perfiles Completada**

✅ Usuarios nuevos: {stats['profiles_created']}
🔄 Perfiles actualizados: {stats['profiles_updated']}
⚙️ Configuraciones: {stats['settings_updated']}
🚫 Límites: {stats['limits_created']}
💜 Kinks: {stats['user_kinks_created']}"""

        if stats["errors"]:
            error_list = "\n".join(stats["errors"][:5])
            if len(stats["errors"]) > 5:
                error_list += f"\n... y {len(stats['errors']) - 5} errores más"
            response += f"\n\n❌ **Errores ({len(stats['errors'])}):**\n{error_list}"

        await update.message.reply_text(response)

        logger.info(
            f"Profile import by {update.effective_user.id}: "
            f"{stats['profiles_created']} created, {stats['profiles_updated']} updated"
        )

    except Exception as e:
        logger.error(f"Profile import error: {e}")
        await update.message.reply_text(f"❌ Error al importar perfiles: {str(e)}")
