"""
The Phantom Bot - Excel Import Handler
/importar - Import users and balances from Excel file
"""
import logging
import tempfile
from pathlib import Path
from telegram import Update
from telegram.ext import ContextTypes

from src.config import settings
from src.database.connection import get_session
from src.database.repositories import UserRepository

logger = logging.getLogger(__name__)


async def importar_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /importar command - show import instructions."""
    if not update.effective_user or not update.message:
        return

    # Check if user is admin
    if update.effective_user.id not in settings.super_admin_ids:
        await update.message.reply_text("❌ Solo los administradores pueden importar datos.")
        return

    await update.message.reply_text(
        """📥 **Importar desde Excel**

Para importar usuarios y saldos:

1. Envía un archivo Excel (.xlsx) con las columnas:
   - username (requerido)
   - balance (requerido)
   - first_name (opcional)

2. El archivo debe tener estas columnas en la primera fila.

3. Envía el archivo como respuesta a este mensaje.

Ejemplo de formato:
| username | balance | first_name |
|----------|---------|------------|
| @usuario1 | 1000   | Juan       |
| @usuario2 | 500    | María      |

⚠️ Los usuarios existentes se actualizarán con el nuevo saldo."""
    )


async def handle_excel_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle Excel document uploads for import."""
    if not update.effective_user or not update.message or not update.message.document:
        return

    # Check if user is admin
    if update.effective_user.id not in settings.super_admin_ids:
        return

    document = update.message.document

    # Check file extension
    if not document.file_name or not document.file_name.endswith(('.xlsx', '.xls')):
        await update.message.reply_text(
            "❌ Solo se aceptan archivos Excel (.xlsx o .xls)"
        )
        return

    try:
        # Import openpyxl here to avoid import errors if not installed
        try:
            import openpyxl
        except ImportError:
            await update.message.reply_text(
                "❌ Error: La librería openpyxl no está instalada.\n"
                "Ejecuta: pip install openpyxl"
            )
            return

        # Download file
        file = await context.bot.get_file(document.file_id)

        # Save to temp file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            await file.download_to_drive(tmp.name)
            tmp_path = Path(tmp.name)

        # Read Excel file
        workbook = openpyxl.load_workbook(tmp_path, read_only=True)
        sheet = workbook.active

        if not sheet:
            await update.message.reply_text("❌ El archivo no tiene hojas de cálculo.")
            return

        # Get headers
        headers = []
        for cell in sheet[1]:
            headers.append(str(cell.value).lower().strip() if cell.value else "")

        # Validate required columns
        if "username" not in headers or "balance" not in headers:
            await update.message.reply_text(
                "❌ El archivo debe tener columnas 'username' y 'balance'"
            )
            return

        username_idx = headers.index("username")
        balance_idx = headers.index("balance")
        first_name_idx = headers.index("first_name") if "first_name" in headers else None

        # Process rows
        imported = 0
        updated = 0
        errors = []

        async with get_session() as session:
            user_repo = UserRepository(session)

            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[username_idx]:
                        continue

                    username = str(row[username_idx]).strip()
                    if username.startswith("@"):
                        username = username[1:]

                    try:
                        balance = int(float(row[balance_idx]))
                    except (ValueError, TypeError):
                        errors.append(f"Fila {row_num}: Balance inválido")
                        continue

                    first_name = None
                    if first_name_idx is not None and row[first_name_idx]:
                        first_name = str(row[first_name_idx]).strip()

                    # Check if user exists
                    existing = await user_repo.get_by_username(username)

                    if existing:
                        # Update balance
                        await user_repo.update_balance(existing.id, balance - existing.balance)
                        updated += 1
                    else:
                        # Create user (without telegram_id, they'll need to /start)
                        await user_repo.create_placeholder(
                            username=username,
                            first_name=first_name,
                            balance=balance,
                        )
                        imported += 1

                except Exception as e:
                    errors.append(f"Fila {row_num}: {str(e)}")
                    continue

        # Clean up temp file
        tmp_path.unlink(missing_ok=True)

        # Build response
        response = f"""📥 **Importación Completada**

✅ Nuevos usuarios: {imported}
🔄 Actualizados: {updated}"""

        if errors:
            error_list = "\n".join(errors[:5])  # Show first 5 errors
            if len(errors) > 5:
                error_list += f"\n... y {len(errors) - 5} errores más"
            response += f"\n\n❌ Errores ({len(errors)}):\n{error_list}"

        await update.message.reply_text(response)

        logger.info(
            f"Excel import by {update.effective_user.id}: "
            f"{imported} imported, {updated} updated, {len(errors)} errors"
        )

    except Exception as e:
        logger.error(f"Excel import error: {e}")
        await update.message.reply_text(f"❌ Error al procesar el archivo: {str(e)}")


async def exportar_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handle /exportar command - export users to Excel."""
    if not update.effective_user or not update.message:
        return

    # Check if user is admin
    if update.effective_user.id not in settings.super_admin_ids:
        await update.message.reply_text("❌ Solo los administradores pueden exportar datos.")
        return

    try:
        # Import openpyxl here
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            await update.message.reply_text(
                "❌ Error: La librería openpyxl no está instalada.\n"
                "Ejecuta: pip install openpyxl"
            )
            return

        async with get_session() as session:
            user_repo = UserRepository(session)
            users = await user_repo.get_all()

            if not users:
                await update.message.reply_text("📊 No hay usuarios para exportar.")
                return

            # Create workbook
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Usuarios"

            # Headers
            headers = ["username", "first_name", "balance", "telegram_id", "is_admin", "created_at"]
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)

            for col, header in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font

            # Data rows
            for row_num, user in enumerate(users, start=2):
                sheet.cell(row=row_num, column=1, value=user.username or "")
                sheet.cell(row=row_num, column=2, value=user.first_name or "")
                sheet.cell(row=row_num, column=3, value=user.balance)
                sheet.cell(row=row_num, column=4, value=user.telegram_id)
                sheet.cell(row=row_num, column=5, value=user.is_admin)
                sheet.cell(row=row_num, column=6, value=user.created_at.isoformat() if user.created_at else "")

            # Adjust column widths
            for col in range(1, len(headers) + 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

            # Save to temp file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                workbook.save(tmp.name)
                tmp_path = Path(tmp.name)

            # Send file
            with open(tmp_path, 'rb') as f:
                await update.message.reply_document(
                    document=f,
                    filename=f"phantom_usuarios_{len(users)}.xlsx",
                    caption=f"📊 Exportación completada: {len(users)} usuarios"
                )

            # Clean up
            tmp_path.unlink(missing_ok=True)

            logger.info(f"Excel export by {update.effective_user.id}: {len(users)} users")

    except Exception as e:
        logger.error(f"Excel export error: {e}")
        await update.message.reply_text(f"❌ Error al exportar: {str(e)}")
